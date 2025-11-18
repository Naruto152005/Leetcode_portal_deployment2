from fastapi import FastAPI, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, FileResponse
import requests
import pandas as pd
import os
from openpyxl.styles import Font, PatternFill
from datetime import datetime

excel_file = "leetcode_recent_solved.xlsx"

query = """
query recentAcSubmissions($username: String!, $limit: Int!) {
  recentAcSubmissionList(username: $username, limit: $limit) {
    id
    title
    titleSlug
    timestamp
    lang
  }
}
"""

# 🔹 Function to get slug → question number mapping (This part is correct)
def get_problem_number_mapping():
    url = "https://leetcode.com/api/problems/all/"
    
    api_headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36"
    }
    
    mapping = {}
    
    try:
        response = requests.get(url, headers=api_headers)
        
        if response.status_code == 200:
            data = response.json()
            for q in data['stat_status_pairs']:
                slug = q['stat']['question__title_slug']
                qid = q['stat']['frontend_question_id']
                mapping[slug] = qid
        else:
            print(f"❌ Failed to fetch problem metadata. Status: {response.status_code}")
            
    except requests.exceptions.RequestException as e:
        print(f"❌ Network error while fetching problem metadata: {e}")

    return mapping

# 🔹 MAIN FUNCTION TO EXPORT (Reverted to the simple, correct version)
def fetch_and_save_recent_problems():
    
    print("Fetching 20 most recent submissions from LeetCode...")
    
    # This is the single, correct API call
    response = requests.post("https://leetcode.com/graphql", json=json_data, headers=headers)
    print("Status code:", response.status_code)

    if response.status_code != 200:
        print("❌ Failed to fetch data from LeetCode")
        print("Response:", response.text) # Added this to show the error
        return

    response_json = response.json()
    
    if 'data' not in response_json or 'recentAcSubmissionList' not in response_json['data']:
        print("❌ Could not find 'recentAcSubmissionList' in API response.")
        print("Response:", response_json)
        return

    # No loop needed, we just get the list
    submissions = response_json['data']['recentAcSubmissionList']
    print(f"Total submissions fetched: {len(submissions)}")
    
    slug_to_qno = get_problem_number_mapping()

    new_data = []
    for sub in submissions:
        sub_id = sub['id'].strip() if isinstance(sub['id'], str) else sub['id']
        slug = sub['titleSlug'].strip()
        
        formatted_date = datetime.fromtimestamp(int(sub['timestamp'])).strftime("%Y-%m-%d %H:%M:%S")

        new_data.append({
            "ID": sub_id,
            "Question No": slug_to_qno.get(slug, "N"
                                             "/A"),
            "Title": sub['title'].strip(),
            "Slug": slug,
            "Timestamp": formatted_date, 
            "Language": sub['lang'].strip(),
            "URL": f"https://leetcode.com/problems/{slug}/"
        })

    new_df = pd.DataFrame(new_data)
    columns_order = ["Question No", "Title", "Language", "URL", "Timestamp", "ID", "Slug"]
    
    if not new_df.empty:
        new_df = new_df[columns_order]

    if os.path.exists(excel_file):
        try:
            old_df = pd.read_excel(excel_file)
        except Exception as e:
            print(f"⚠️ Warning: Could not read existing Excel file. Starting fresh. Error: {e}")
            old_df = pd.DataFrame()
    else:
        old_df = pd.DataFrame()

    print(f"Old data count: {len(old_df)}")
    print(f"New data count: {len(new_df)}")

    combined_before = pd.concat([old_df, new_df])
    print(f"Combined before removing duplicates: {len(combined_before)}")

    # This line correctly handles duplicates by keeping the newest data
    combined_df = combined_before.drop_duplicates(subset=["Slug"], keep='last').reset_index(drop=True)

    print(f"Total unique solved problems: {len(combined_df)}")

    # === Save and Style Block (This part is correct) ===
    
    if combined_df.empty:
        print("ℹ️ No data to save. Excel file not modified.")
        return 

    try:
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            combined_df.to_excel(writer, index=False, sheet_name='RecentSubmissions')

            ws = writer.sheets['RecentSubmissions']

            header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            header_font = Font(bold=True)

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font

            column_widths = {
                "A": 14, "B": 35, "C": 15, "D": 50, "E": 20, "F": 25, "G": 30
            }
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width

        print(f"✅ Data saved to {excel_file}")
        print("🎨 Styles applied to Excel.")

    except Exception as e:
        print(f"❌ Error during Excel saving/styling: {e}")
        print("Saving data without styles as a fallback...")
        combined_df.to_excel(excel_file, index=False)
        
app = FastAPI()
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.post("/api/recent-submissions")
def api_recent_submissions(username: str = Form(...), session: str = Form(...), csrf: str = Form(...)):
    headers = {
        "Content-Type": "application/json",
        "x-csrftoken": csrf,
        "Referer": "https://leetcode.com",
        "User-Agent": "Mozilla/5.0",
    }
    cookies = {"LEETCODE_SESSION": session, "csrftoken": csrf}
    variables = {"username": username, "limit": 20}
    json_data = {"query": query, "variables": variables}
    r = requests.post("https://leetcode.com/graphql", json=json_data, headers=headers, cookies=cookies)
    d = r.json()
    if "data" not in d or "recentAcSubmissionList" not in d["data"]:
        return JSONResponse({"error": "Invalid credentials or API error"}, status_code=400)
    subs = d["data"]["recentAcSubmissionList"]
    enriched = [
        {
            "id": s.get("id"),
            "title": s.get("title"),
            "titleSlug": s.get("titleSlug"),
            "link": f"https://leetcode.com/problems/{s.get('titleSlug')}/",
            "date": datetime.fromtimestamp(int(s.get("timestamp"))).strftime("%Y-%m-%d %H:%M:%S"),
            "lang": s.get("lang"),
        }
        for s in subs
    ]
    return JSONResponse(enriched)

@app.post("/api/download-excel")
def api_download_excel(username: str = Form(...), session: str = Form(...), csrf: str = Form(...)):
    headers = {
        "Content-Type": "application/json",
        "x-csrftoken": csrf,
        "Referer": "https://leetcode.com",
        "User-Agent": "Mozilla/5.0",
    }
    cookies = {"LEETCODE_SESSION": session, "csrftoken": csrf}
    variables = {"username": username, "limit": 20}
    json_data = {"query": query, "variables": variables}
    r = requests.post("https://leetcode.com/graphql", json=json_data, headers=headers, cookies=cookies)
    d = r.json()
    if "data" not in d or "recentAcSubmissionList" not in d["data"]:
        return JSONResponse({"error": "Invalid credentials or API error"}, status_code=400)
    subs = d["data"]["recentAcSubmissionList"]
    slug_to_qno = get_problem_number_mapping()
    rows = [
        {
            "Question No": slug_to_qno.get(s.get("titleSlug"), ""),
            "Title": s.get("title"),
            "Language": s.get("lang"),
            "URL": f"https://leetcode.com/problems/{s.get('titleSlug')}/",
            "Timestamp": datetime.fromtimestamp(int(s.get("timestamp"))).strftime("%Y-%m-%d %H:%M:%S"),
            "ID": s.get("id"),
            "Slug": s.get("titleSlug"),
        }
        for s in subs
    ]
    df = pd.DataFrame(rows)
    file_path = f"{username}_submissions.xlsx"
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="RecentSubmissions")
        ws = writer.sheets["RecentSubmissions"]
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        column_widths = {"A": 14, "B": 35, "C": 15, "D": 50, "E": 20, "F": 25, "G": 30}
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
    return FileResponse(file_path, filename=file_path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    